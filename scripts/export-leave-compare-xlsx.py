# -*- coding: utf-8 -*-
"""직원별 연차: DB 기준 vs 입사일 기준 비교 엑셀 → 바탕화면"""
from __future__ import annotations

import json
import re
import subprocess
import sys
from datetime import date, datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
TODAY = date(2026, 7, 15)


# 운영 DB 이름. wrangler.toml 의 database_name 과 일치해야 한다.
# 구 DB 'pchos-d1' 을 가리키던 동안에는 조회 결과가 비어 리포트가 조용히 빈 채로 나왔다.
D1_DB_NAME = "pchos-d1-v2"


def d1(sql: str) -> list[dict]:
    r = subprocess.run(
        ["npx", "wrangler", "d1", "execute", D1_DB_NAME, "--remote", "--json", "--command", sql],
        capture_output=True,
        cwd=str(ROOT),
        shell=True,
    )
    out = (r.stdout or b"").decode("utf-8", errors="replace")
    err = (r.stderr or b"").decode("utf-8", errors="replace")
    if r.returncode != 0:
        print(err[-800:], file=sys.stderr)
        raise RuntimeError(f"d1 failed: {r.returncode}")
    m = re.search(r"(\[\s*\{|\{\s*\")", out)
    if not m:
        raise RuntimeError("no JSON in wrangler output")
    data = json.loads(out[m.start() :])
    if isinstance(data, list):
        for block in data:
            if isinstance(block, dict) and "results" in block:
                return block["results"] or []
        return data
    if isinstance(data, dict) and "results" in data:
        return data["results"] or []
    return []


def parse_ymd(s) -> date | None:
    if not s:
        return None
    m = re.match(r"(\d{4})-(\d{2})-(\d{2})", str(s).strip())
    if not m:
        return None
    try:
        return date(int(m.group(1)), int(m.group(2)), int(m.group(3)))
    except ValueError:
        return None


def days_in_month(y: int, m: int) -> int:
    if m == 12:
        return 31
    return (date(y, m + 1, 1) - date(y, m, 1)).days


def add_months(d: date, months: int) -> date:
    total = d.year * 12 + (d.month - 1) + months
    y = total // 12
    m = total % 12 + 1
    return date(y, m, min(d.day, days_in_month(y, m)))


def add_years(d: date, years: int) -> date:
    y = d.year + years
    return date(y, d.month, min(d.day, days_in_month(y, d.month)))


def tenure_years(hire: date, today: date) -> int:
    years = today.year - hire.year
    if (today.month, today.day) < (hire.month, hire.day):
        years -= 1
    return max(0, years)


def annual_days_for_tenure(n: int) -> float:
    if n < 1:
        return 0.0
    return float(min(25, 15 + (n - 1) // 2))


def is_annual_use(leave_type: str) -> bool:
    t = (leave_type or "").strip()
    if not t or "부여" in t:
        return False
    return "연차" in t or t.lower() in ("annual", "annual_leave")


def is_half(leave_type: str) -> bool:
    t = (leave_type or "").strip()
    return "반차" in t


def leave_days(row: dict) -> float:
    lt = str(row.get("leave_type") or "")
    if is_half(lt):
        return 0.5
    if row.get("days") is not None:
        try:
            d = float(row["days"])
            if d > 0:
                return d
        except (TypeError, ValueError):
            pass
    start = parse_ymd(row.get("start_date"))
    end = parse_ymd(row.get("end_date") or row.get("start_date"))
    if not start:
        return 0.0
    if not end or end < start:
        return 1.0
    return float((end - start).days + 1)


def resolve_hire(s: dict) -> date | None:
    return parse_ymd(s.get("hire_date")) or parse_ymd(s.get("join_date")) or parse_ymd(s.get("joined_at"))


def desktop_path() -> Path:
    home = Path.home()
    for c in [
        home / "OneDrive" / "Desktop",
        home / "OneDrive" / "바탕 화면",
        home / "Desktop",
        home / "바탕 화면",
    ]:
        if c.is_dir():
            return c
    try:
        r = subprocess.run(
            ["powershell", "-NoProfile", "-Command", "[Environment]::GetFolderPath('Desktop')"],
            capture_output=True,
            text=True,
        )
        p = Path((r.stdout or "").strip())
        if p.is_dir():
            return p
    except Exception:
        pass
    return home


def n(v) -> float:
    if v is None or v == "":
        return 0.0
    try:
        return float(v)
    except (TypeError, ValueError):
        return 0.0


def main():
    print("D1 조회…")
    staff = d1(
        "SELECT s.id, s.name, s.employee_no, s.company, s.department, s.team, s.status, "
        "s.hire_date, s.join_date, s.joined_at, s.annual_leave_total, s.annual_leave_used, "
        "b.total_days, b.used_days, b.remaining_days, b.expired_days, b.compensated_days, "
        "b.expiry_date "
        "FROM staff_members s "
        "LEFT JOIN leave_balances b ON b.staff_id = s.id AND b.year = 2026 "
        "ORDER BY COALESCE(s.hire_date, s.join_date, s.joined_at), s.name"
    )
    accruals = d1(
        "SELECT staff_id, kind, period_key, days, source_date, note FROM leave_accruals "
        "ORDER BY staff_id, kind, period_key"
    )
    requests = d1(
        "SELECT staff_id, leave_type, start_date, end_date, days FROM leave_requests "
        "WHERE status IN ('승인', 'approved') ORDER BY staff_id, start_date"
    )
    print(f"staff={len(staff)} accruals={len(accruals)} req={len(requests)}")

    by_acc: dict[str, list] = {}
    for a in accruals:
        by_acc.setdefault(str(a["staff_id"]), []).append(a)
    by_req: dict[str, list] = {}
    for r in requests:
        by_req.setdefault(str(r["staff_id"]), []).append(r)

    thin = Border(
        left=Side(style="thin", color="BFBFBF"),
        right=Side(style="thin", color="BFBFBF"),
        top=Side(style="thin", color="BFBFBF"),
        bottom=Side(style="thin", color="BFBFBF"),
    )
    font = Font(name="Arial", size=10)
    font_b = Font(name="Arial", size=10, bold=True)
    font_h = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    font_title = Font(name="Arial", size=14, bold=True, color="1F4E79")
    fill_base = PatternFill("solid", fgColor="1F4E79")
    fill_db = PatternFill("solid", fgColor="2E75B6")
    fill_hire = PatternFill("solid", fgColor="548235")
    fill_cmp = PatternFill("solid", fgColor="C65911")
    fill_db_row = PatternFill("solid", fgColor="DDEBF7")
    fill_hire_row = PatternFill("solid", fgColor="E2EFDA")
    fill_bad = PatternFill("solid", fgColor="FCE4D6")
    fill_ok = PatternFill("solid", fgColor="C6EFCE")
    align_c = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_l = Alignment(horizontal="left", vertical="center")

    wb = Workbook()
    ws = wb.active
    ws.title = "직원별_DB_입사일비교"

    # Row 1: group headers
    # A-G 기본, H-K DB, L-P 입사일, Q-S 비교
    headers_group = [
        (1, 7, "기본정보", fill_base),
        (8, 12, "【DB 기준】 leave_balances 2026", fill_db),
        (13, 18, "【입사일 기준】 근로기준법 발생·사용", fill_hire),
        (19, 22, "차이 비교", fill_cmp),
    ]
    for c1, c2, title, fill in headers_group:
        ws.merge_cells(start_row=1, start_column=c1, end_row=1, end_column=c2)
        cell = ws.cell(row=1, column=c1, value=title)
        cell.fill = fill
        cell.font = font_h
        cell.alignment = align_c
        for c in range(c1, c2 + 1):
            ws.cell(row=1, column=c).fill = fill
            ws.cell(row=1, column=c).border = thin

    col_headers = [
        "이름", "사번", "회사", "부서", "상태", "입사일", "근속(만년)",
        # DB
        "DB_총부여", "DB_사용", "DB_잔여", "DB_소멸", "DB_만료일",
        # Hire
        "입사_총발생", "입사_발생구분", "입사_사용", "입사_잔여", "입사_사용기간", "입사_연차키",
        # Compare
        "총부여차(DB-입사)", "사용차(DB-입사)", "잔여차(DB-입사)", "판정",
    ]
    for c, h in enumerate(col_headers, 1):
        cell = ws.cell(row=2, column=c, value=h)
        cell.font = font_h
        cell.alignment = align_c
        cell.border = thin
        if c <= 7:
            cell.fill = fill_base
        elif c <= 12:
            cell.fill = fill_db
        elif c <= 18:
            cell.fill = fill_hire
        else:
            cell.fill = fill_cmp
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 28

    data_rows = []
    for s in staff:
        sid = str(s["id"])
        status = (s.get("status") or "").strip()
        hire = resolve_hire(s)
        hire_str = hire.isoformat() if hire else ""
        ty = tenure_years(hire, TODAY) if hire else None

        # DB
        db_total = s.get("total_days")
        db_used = s.get("used_days")
        db_remain = s.get("remaining_days")
        db_expired = n(s.get("expired_days"))
        db_comp = n(s.get("compensated_days"))
        db_expiry = s.get("expiry_date") or ""
        # fallback staff fields if no balance row
        if db_total is None:
            db_total = s.get("annual_leave_total")
        if db_used is None:
            db_used = s.get("annual_leave_used")
        if db_remain is None and db_total is not None and db_used is not None:
            db_remain = max(0.0, n(db_total) - n(db_used) - db_expired - db_comp)

        db_total_f = n(db_total) if db_total is not None else None
        db_used_f = n(db_used) if db_used is not None else None
        db_remain_f = n(db_remain) if db_remain is not None else None

        # Accruals for hire-based grant
        accs = by_acc.get(sid, [])
        monthly = [a for a in accs if a.get("kind") == "monthly"]
        annuals = [a for a in accs if a.get("kind") == "annual"]
        monthly_sum = sum(n(a.get("days")) for a in monthly)

        latest_n = 0
        latest_days = 0.0
        latest_key = ""
        for a in annuals:
            nn = int(re.sub(r"^annual:", "", str(a.get("period_key") or "0")) or 0)
            if nn >= latest_n:
                latest_n = nn
                latest_days = n(a.get("days"))
                latest_key = str(a.get("period_key") or "")

        # Hire-based expected grant (policy)
        hire_grant = 0.0
        hire_kind = ""
        hire_keys = ""
        if hire:
            ty_v = tenure_years(hire, TODAY)
            if ty_v >= 1:
                # 입사일 기준 발생: 만 N년차 연차 일수 (최신 N)
                hire_grant = annual_days_for_tenure(ty_v)
                hire_kind = f"연차 만{ty_v}년차"
                hire_keys = f"annual:{ty_v}"
                # DB 원장에 있으면 그 일수 우선 (부여 기록 반영)
                if latest_n >= ty_v and latest_days > 0:
                    hire_grant = latest_days
                    hire_keys = latest_key or hire_keys
                elif latest_days > 0 and latest_n > 0:
                    # 원장 최신 연차 사용
                    hire_grant = latest_days
                    hire_keys = latest_key
                    hire_kind = f"연차 {latest_key}"
            else:
                # 1년 미만: 경과 월수(최대 11)
                exp_m = 0
                for k in range(1, 12):
                    if add_months(hire, k) <= TODAY:
                        exp_m += 1
                # 실제 월차 부여 건수와 기대 중 실제 원장 우선 표시, 없으면 기대
                actual_m = len(monthly)
                hire_grant = float(actual_m if actual_m > 0 else exp_m)
                if monthly_sum > 0:
                    hire_grant = monthly_sum
                hire_kind = f"월차 {int(hire_grant)}개월"
                hire_keys = f"monthly×{int(hire_grant)}"
        else:
            hire_kind = "입사일없음"

        # Hire-based usage: within current leave period
        # 1년 이상: 최근 응당일(만 N년차 시작일) ~ 다음 응당일
        # 1년 미만: 입사일 ~ 오늘 (또는 1년 응당일 전)
        hire_used = 0.0
        period_label = ""
        if hire:
            ty_v = tenure_years(hire, TODAY)
            if ty_v >= 1:
                period_start = add_years(hire, ty_v)
                period_end = add_years(hire, ty_v + 1)
                period_label = f"{period_start.isoformat()} ~ {period_end.isoformat()}"
            else:
                period_start = hire
                period_end = add_years(hire, 1)
                period_label = f"{period_start.isoformat()} ~ {period_end.isoformat()} (1년미만)"
            for r in by_req.get(sid, []):
                if not is_annual_use(str(r.get("leave_type") or "")):
                    continue
                sd = parse_ymd(r.get("start_date"))
                if not sd:
                    continue
                if period_start <= sd < period_end:
                    hire_used += leave_days(r)
        else:
            period_label = "-"

        hire_remain = max(0.0, round(hire_grant - hire_used, 2))

        # Diffs
        d_total = (db_total_f - hire_grant) if db_total_f is not None else None
        d_used = (db_used_f - hire_used) if db_used_f is not None else None
        d_remain = (db_remain_f - hire_remain) if db_remain_f is not None else None

        judge = "일치"
        if d_total is not None and abs(d_total) > 0.01:
            judge = "총부여불일치"
        if d_used is not None and abs(d_used) > 0.01:
            judge = "사용불일치" if judge == "일치" else judge + "+사용"
        if d_remain is not None and abs(d_remain) > 0.01:
            judge = "잔여불일치" if judge == "일치" else judge + "+잔여"
        if status and status not in ("재직", "재직중", "active", "Active", ""):
            judge = f"비재직/{judge}" if judge != "일치" else "비재직"
        if not hire:
            judge = "입사일없음"

        row = [
            s.get("name") or "",
            s.get("employee_no") or "",
            s.get("company") or "",
            s.get("department") or "",
            status,
            hire_str,
            ty if ty is not None else "",
            db_total_f if db_total_f is not None else "",
            db_used_f if db_used_f is not None else "",
            db_remain_f if db_remain_f is not None else "",
            db_expired if db_expired else 0,
            db_expiry,
            hire_grant,
            hire_kind,
            hire_used,
            hire_remain,
            period_label,
            hire_keys,
            round(d_total, 2) if d_total is not None else "",
            round(d_used, 2) if d_used is not None else "",
            round(d_remain, 2) if d_remain is not None else "",
            judge,
        ]
        data_rows.append(row)

    for i, row in enumerate(data_rows, start=3):
        for c, val in enumerate(row, 1):
            cell = ws.cell(row=i, column=c, value=val if val != "" else None)
            cell.font = font
            cell.border = thin
            cell.alignment = align_c if c >= 7 else align_l
            if 8 <= c <= 12:
                cell.fill = fill_db_row
            elif 13 <= c <= 18:
                cell.fill = fill_hire_row
        # judge coloring
        jcell = ws.cell(row=i, column=22)
        if jcell.value == "일치":
            jcell.fill = fill_ok
            jcell.font = font_b
        elif jcell.value and jcell.value != "비재직":
            jcell.fill = fill_bad
            jcell.font = font_b

    last = 2 + len(data_rows)
    ws.auto_filter.ref = f"A2:V{last}"
    ws.freeze_panes = "C3"

    widths = {
        "A": 10, "B": 8, "C": 16, "D": 10, "E": 8, "F": 12, "G": 10,
        "H": 10, "I": 10, "J": 10, "K": 9, "L": 12,
        "M": 11, "N": 14, "O": 10, "P": 10, "Q": 28, "R": 12,
        "S": 12, "T": 12, "U": 12, "V": 16,
    }
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    # 안내 sheet
    ws0 = wb.create_sheet("안내", 0)
    lines = [
        ["직원별 연차 비교 — DB 기준 vs 입사일 기준", ""],
        ["생성일시", datetime.now().strftime("%Y-%m-%d %H:%M")],
        ["기준일", TODAY.isoformat()],
        ["DB", f"Cloudflare D1 {D1_DB_NAME}"],
        ["", ""],
        ["【DB 기준】", "leave_balances (year=2026) 저장값"],
        ["DB_총부여", "total_days (없으면 staff.annual_leave_total)"],
        ["DB_사용", "used_days — 주로 달력 2026년 승인 연차 합"],
        ["DB_잔여", "remaining_days = 총−사용−소멸−보상"],
        ["", ""],
        ["【입사일 기준】", "hire_date 기준 근로기준법 구간"],
        ["입사_총발생", "1년 미만: 경과 월차 일수 / 1년 이상: 만 N년차 연차 일수(15+…)"],
        ["입사_사용", "현재 연차 사용기간(최근 응당일~다음 응당일) 안 승인 사용만 합산"],
        ["입사_잔여", "입사_총발생 − 입사_사용"],
        ["입사_사용기간", "입사 응당일 기준 당해 연차 구간"],
        ["", ""],
        ["차이", "DB값 − 입사일기준값 (0이면 일치)"],
        ["", ""],
        ["김영대 예시", "DB 15/7/8 (달력2026 사용) vs 입사 15/3/12 (2026-03-17~2027-03-17 구간 사용)"],
    ]
    for r in lines:
        ws0.append(r)
    ws0["A1"].font = font_title
    ws0.column_dimensions["A"].width = 18
    ws0.column_dimensions["B"].width = 80

    # 재직자만 시트
    ws2 = wb.create_sheet("재직자만")
    for c1, c2, title, fill in headers_group:
        ws2.merge_cells(start_row=1, start_column=c1, end_row=1, end_column=c2)
        cell = ws2.cell(row=1, column=c1, value=title)
        cell.fill = fill
        cell.font = font_h
        cell.alignment = align_c
        for c in range(c1, c2 + 1):
            ws2.cell(row=1, column=c).fill = fill
            ws2.cell(row=1, column=c).border = thin
    for c, h in enumerate(col_headers, 1):
        cell = ws2.cell(row=2, column=c, value=h)
        cell.font = font_h
        cell.alignment = align_c
        cell.border = thin
        if c <= 7:
            cell.fill = fill_base
        elif c <= 12:
            cell.fill = fill_db
        elif c <= 18:
            cell.fill = fill_hire
        else:
            cell.fill = fill_cmp
    ws2.row_dimensions[1].height = 22
    ws2.row_dimensions[2].height = 28
    ri = 3
    for row in data_rows:
        st = str(row[4] or "")
        if st and st not in ("재직", "재직중", "active", "Active"):
            continue
        for c, val in enumerate(row, 1):
            cell = ws2.cell(row=ri, column=c, value=val if val != "" else None)
            cell.font = font
            cell.border = thin
            cell.alignment = align_c if c >= 7 else align_l
            if 8 <= c <= 12:
                cell.fill = fill_db_row
            elif 13 <= c <= 18:
                cell.fill = fill_hire_row
        jcell = ws2.cell(row=ri, column=22)
        if jcell.value == "일치":
            jcell.fill = fill_ok
            jcell.font = font_b
        elif jcell.value:
            jcell.fill = fill_bad
            jcell.font = font_b
        ri += 1
    ws2.auto_filter.ref = f"A2:V{max(2, ri - 1)}"
    ws2.freeze_panes = "C3"
    for col, w in widths.items():
        ws2.column_dimensions[col].width = w

    out = desktop_path() / f"연차_DB_입사일_비교_{TODAY.isoformat()}.xlsx"
    wb.save(out)
    print(f"saved: {out}")

    # preview sample
    for row in data_rows:
        if row[0] in ("김영대", "김이지", "이미영", "조숙현"):
            print(
                f"{row[0]} | DB {row[7]}/{row[8]}/{row[9]} | "
                f"입사 {row[12]}/{row[14]}/{row[15]} ({row[13]}) | {row[21]}"
            )


if __name__ == "__main__":
    main()
