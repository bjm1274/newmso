# -*- coding: utf-8 -*-
"""연차관리대장 엑셀 파싱 + Supabase 조회 공통 모듈.
preview.py / apply.py / revert.py 가 공유한다."""
import os
import re
import sys
import io
import datetime
import json
import urllib.request
import urllib.parse

if sys.stdout.encoding and sys.stdout.encoding.lower() != "utf-8":
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

EXCEL_PATH = r"C:\Users\baek_\OneDrive\바탕 화면\연차관리대장 (1).xlsx"
TODAY = datetime.date(2026, 5, 14)
COMMAND_TS = "2026-05-14T16:54:00"  # 회귀 기준 시각 (명령 입력 시각)
SKIP_SHEETS = {"연차대장_전체", "신규", "유민우", "김대운", "최은아"}
PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
ENV_PATH = os.path.join(PROJECT_ROOT, ".env.local")


# ---------------------------------------------------------------- env / supabase
def load_env():
    env = {}
    with open(ENV_PATH, "r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line or line.startswith("#") or "=" not in line:
                continue
            k, v = line.split("=", 1)
            env[k.strip()] = v.strip().strip('"').strip("'")
    return env


_ENV = None


def _env():
    global _ENV
    if _ENV is None:
        _ENV = load_env()
    return _ENV


def sb_request(method, path, body=None, params=None, prefer=None):
    """Supabase REST 호출 (service role key)."""
    env = _env()
    url = env["NEXT_PUBLIC_SUPABASE_URL"].rstrip("/") + "/rest/v1/" + path
    if params:
        url += "?" + urllib.parse.urlencode(params, safe="(),.*:")
    key = env["SUPABASE_SERVICE_ROLE_KEY"]
    pref = "return=representation"
    if prefer:
        pref += "," + prefer
    headers = {
        "apikey": key,
        "Authorization": f"Bearer {key}",
        "Content-Type": "application/json",
        "Prefer": pref,
    }
    data = json.dumps(body).encode("utf-8") if body is not None else None
    req = urllib.request.Request(url, data=data, headers=headers, method=method)
    try:
        with urllib.request.urlopen(req) as resp:
            raw = resp.read().decode("utf-8")
            return json.loads(raw) if raw else []
    except urllib.error.HTTPError as e:
        detail = e.read().decode("utf-8", "replace")
        raise RuntimeError(f"Supabase {method} {path} 실패 ({e.code}): {detail}")


def fetch_staff_members():
    cols = "id,name,employee_no,company,company_id,join_date,joined_at,hire_date,annual_leave_total,annual_leave_used,annual_leave_pay,status"
    return sb_request("GET", "staff_members", params={"select": cols})


def fetch_leave_balances(year):
    return sb_request("GET", "leave_balances", params={"select": "*", "year": f"eq.{year}"})


def fetch_leave_requests(staff_ids):
    if not staff_ids:
        return []
    out = []
    CHUNK = 40
    for i in range(0, len(staff_ids), CHUNK):
        chunk = staff_ids[i:i + CHUNK]
        ids = ",".join(chunk)
        out += sb_request(
            "GET", "leave_requests",
            params={"select": "id,staff_id,leave_type,start_date,end_date,status,reason,approved_at,created_at",
                    "staff_id": f"in.({ids})"},
        )
    return out


# ---------------------------------------------------------------- excel parsing
DATE_MARKERS_SKIP = ("수당지급", "소멸")
NON_ANNUAL_MARKERS = ("병가", "군소집", "경조", "대체", "조퇴", "반차 ", "반반차")


def _to_date(v):
    if isinstance(v, datetime.datetime):
        return v.date()
    if isinstance(v, datetime.date):
        return v
    if isinstance(v, str):
        s = v.strip()
        m = re.match(r"^(\d{4})[.\-/](\d{1,2})[.\-/](\d{1,2})$", s)
        if m:
            return datetime.date(int(m[1]), int(m[2]), int(m[3]))
    return None


def parse_date_token(tok, default_year=None):
    """'24.06.28', '2024.06.28', '26.01.18' → date. M/D 형식은 default_year 필요."""
    tok = tok.strip()
    m = re.match(r"^(\d{2,4})[.\-](\d{1,2})[.\-](\d{1,2})$", tok)
    if m:
        y = int(m[1])
        if y < 100:
            y += 2000
        try:
            return datetime.date(y, int(m[2]), int(m[3]))
        except ValueError:
            return None
    m = re.match(r"^(\d{1,2})/(\d{1,2})$", tok)
    if m and default_year:
        try:
            return datetime.date(default_year, int(m[1]), int(m[2]))
        except ValueError:
            return None
    return None


def classify_cell(raw):
    """엑셀 사용일 셀 1개를 분류.
    returns (kind, dates, note)
      kind: 'date' | 'marker' | 'non_annual' | 'garbage'
    """
    if raw is None:
        return ("garbage", [], "")
    s = str(raw).strip()
    if not s:
        return ("garbage", [], "")
    # 순수 마커
    if s in DATE_MARKERS_SKIP:
        return ("marker", [], s)
    # 비-연차 메모 (병가/대체/조퇴/경조/군소집/반차메모)
    if any(mk in s for mk in NON_ANNUAL_MARKERS):
        # 단, 'YY.MM.DD' 로 시작하면서 뒤에 (...) 대체 형태 → 비연차
        return ("non_annual", [], s)
    # 깨진 시트 잡데이터
    if re.search(r"(월$|월별|연차휴가|^\d+일\(|^O$|점심|출근|연장|^대성$|^소연$|^가영$)", s):
        return ("garbage", [], s)
    # 순수 날짜
    d = parse_date_token(s)
    if d:
        return ("date", [d], "")
    # 'M/D, M/D' 다중 (깨진 시트 잔재 가능) — 일단 garbage 로 두고 flag
    return ("garbage", [], s)


def parse_excel():
    """엑셀 → 직원별 구조화 데이터 list."""
    import openpyxl
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    result = []
    for name in wb.sheetnames:
        if name in SKIP_SHEETS:
            continue
        ws = wb[name]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 9:
            result.append({"sheet": name, "valid": False, "reason": "행 부족"})
            continue
        info = rows[3]
        dept = info[1]
        sheet_name = info[2]
        join_date = _to_date(info[3])
        excel_remaining = info[4]
        flags = []
        if join_date is None or sheet_name in (None, "신규") or str(dept) == "#N/A":
            result.append({"sheet": name, "valid": False, "reason": "입사일/이름 #N/A"})
            continue

        # --- 데이터 행 1차 수집 ---
        data_rows = []   # (ym, 발생, 사용, 소멸, cells, 수당지급여부)
        for r in rows[8:]:
            ym = _to_date(r[1])
            if ym is None:
                continue
            g = r[2] if isinstance(r[2], (int, float)) else None
            u = r[3] if isinstance(r[3], (int, float)) else None
            e = r[4] if isinstance(r[4], (int, float)) else None
            cells = r[6:]
            is_comp = any("수당지급" in str(c) for c in cells if c)
            data_rows.append((ym, g, u, e, cells, is_comp))

        # --- 현재 사이클 시작 = TODAY 이전 마지막 '발생>=11(연 15개 부여)' 행 ---
        cstart = data_rows[0][0] if data_rows else join_date
        for ym, g, u, e, cells, is_comp in data_rows:
            if g is not None and g >= 11 and ym <= TODAY:
                cstart = ym

        grant = used = expired = compensated = 0.0
        all_dates = []
        flagged_cells = []
        used_count_mismatch = []

        for ym, g, u, e, cells, is_comp in data_rows:
            in_cycle = cstart <= ym <= TODAY
            if in_cycle:
                if g:
                    grant += g
                if u:
                    used += u
                if e:
                    if is_comp:
                        compensated += e
                    else:
                        expired += e
            row_dates = []
            for c in cells:
                kind, dates, note = classify_cell(c)
                if kind == "date":
                    row_dates += dates
                elif kind in ("non_annual", "garbage") and note:
                    flagged_cells.append((ym.isoformat(), note))
            for d in row_dates:
                all_dates.append({"date": d.isoformat(), "row_ym": ym.isoformat(),
                                  "in_cycle": d >= cstart})
            if u is not None and u > 0 and len(row_dates) != u:
                used_count_mismatch.append({"row_ym": ym.isoformat(), "사용": u,
                                            "추출날짜수": len(row_dates),
                                            "dates": [d.isoformat() for d in row_dates]})

        remaining_calc = round(grant - used - expired - compensated, 2)
        if excel_remaining is not None and isinstance(excel_remaining, (int, float)):
            if abs(remaining_calc - excel_remaining) > 0.01:
                flags.append(f"시트표기 잔여({excel_remaining}) ≠ 계산 잔여({remaining_calc})")

        result.append({
            "sheet": name,
            "valid": True,
            "name": sheet_name,
            "dept": dept,
            "join_date": join_date.isoformat(),
            "cycle_start": cstart.isoformat(),
            "excel_remaining_header": excel_remaining,
            "grant_current_cycle": round(grant, 2),
            "used_current_cycle": round(used, 2),
            "expired_current_cycle": round(expired, 2),
            "compensated_current_cycle": round(compensated, 2),
            "remaining_calc": remaining_calc,
            "usage_dates": all_dates,
            "usage_dates_count": len(all_dates),
            "flagged_cells": flagged_cells,
            "used_count_mismatch": used_count_mismatch,
            "flags": flags,
        })
    return result


def norm_name(s):
    return re.sub(r"\s+", "", str(s or "")).strip()


# ---------------------------------------------------------------- 반영 대상/제외
# 엑셀 자체 불일치 또는 시트 손상으로 이번 반영에서 제외하는 직원 (사용자 결정)
EXCLUDE_NAMES = {"박지영", "최찬", "박하연", "김지은", "이대성"}
EXCLUDE_REASON = {
    "박지영": "엑셀 불일치 — 시트표기 잔여 39 / 개요표 0 / 행 계산 15 (세 값이 전부 다름)",
    "최찬": "엑셀 불일치 — 시트표기 잔여 14 / 행 계산 15 (이전 사이클 초과사용분 처리 차이)",
    "박하연": "엑셀 불일치 — 시트표기 잔여 6 / 행 계산 9",
    "김지은": "시트 손상 — 사용일 셀에 '1월','연차휴가' 등 잡데이터만, 날짜 추출 불가",
    "이대성": "시트 손상 — 옆 표 데이터(O/점심30분/대성·소연·가영) 혼입",
}

LEAVE_TYPE_CURRENT = "연차"
LEAVE_TYPE_PAST = "연차(이력)"  # isAnnualLeaveType 에서 '이력' 포함 시 비집계 처리됨
IMPORT_REASON = "연차관리대장 일괄 반영(2026-05-14)"


def compute_expiry_date(hire_date_str, today=TODAY):
    """calculateAnnualLeaveExpiryDate 와 동일: 올해 입사기념일, 지났으면 내년."""
    if not hire_date_str:
        return datetime.date(today.year, 12, 31)
    hd = _to_date(hire_date_str)
    if hd is None:
        return datetime.date(today.year, 12, 31)
    try:
        exp = datetime.date(today.year, hd.month, hd.day)
    except ValueError:
        exp = datetime.date(today.year, hd.month, 28)
    if exp <= today:
        try:
            exp = datetime.date(today.year + 1, hd.month, hd.day)
        except ValueError:
            exp = datetime.date(today.year + 1, hd.month, 28)
    return exp


def split_usage_dates(emp):
    """엑셀 사용일을 현재 사이클 / 과거로 분리. 반환 (current[], past[]).
    사용일이 어느 사이클에 차감됐는지는 '날짜'가 아니라 그 사용일이 적힌
    '발생월 행(row_ym)' 으로 결정된다 — 행이 사이클 시작 이후면 현재 사이클."""
    cstart = emp["cycle_start"]
    current, past = [], []
    for d in emp["usage_dates"]:
        (current if d["row_ym"] >= cstart else past).append(d["date"])
    return sorted(current), sorted(past)


OUT_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "산출물")
os.makedirs(OUT_DIR, exist_ok=True)
BACKUP_PATH = os.path.join(OUT_DIR, "백업_20260514_1654.json")
