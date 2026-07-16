# -*- coding: utf-8 -*-
"""전 직원 연차 DB vs 입사일 기준 발생/사용 감사 엑셀 생성 → 바탕화면"""
from __future__ import annotations

import json
import re
import subprocess
import sys
from datetime import date, datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
TODAY = date(2026, 7, 15)  # KST 기준 작업일 (세션 기준일)


def d1(sql: str) -> list[dict]:
    r = subprocess.run(
        ["npx", "wrangler", "d1", "execute", "pchos-d1", "--remote", "--json", "--command", sql],
        capture_output=True,
        cwd=str(ROOT),
        shell=True,
    )
    out = (r.stdout or b"").decode("utf-8", errors="replace")
    err = (r.stderr or b"").decode("utf-8", errors="replace")
    if r.returncode != 0:
        print(err[-800:], file=sys.stderr)
        raise RuntimeError(f"d1 failed: {r.returncode}")
    # wrangler prints log lines then JSON
    m = re.search(r"(\[\s*\{|\{\s*\")", out)
    if not m:
        raise RuntimeError("no JSON in wrangler output")
    data = json.loads(out[m.start() :])
    if isinstance(data, list):
        for block in data:
            if isinstance(block, dict) and "results" in block:
                return block["results"] or []
        return data
    if isinstance(data, dict) and "results" in data:
        return data["results"] or []
    return []


def parse_ymd(s: str | None) -> date | None:
    if not s:
        return None
    m = re.match(r"(\d{4})-(\d{2})-(\d{2})", str(s).strip())
    if not m:
        return None
    try:
        return date(int(m.group(1)), int(m.group(2)), int(m.group(3)))
    except ValueError:
        return None


def days_in_month(y: int, m: int) -> int:
    if m == 12:
        return 31
    return (date(y, m + 1, 1) - date(y, m, 1)).days


def add_months(d: date, months: int) -> date:
    total = d.year * 12 + (d.month - 1) + months
    y = total // 12
    m = total % 12 + 1
    day = min(d.day, days_in_month(y, m))
    return date(y, m, day)


def add_years(d: date, years: int) -> date:
    y = d.year + years
    day = min(d.day, days_in_month(y, d.month))
    return date(y, d.month, day)


def tenure_years(hire: date, today: date) -> int:
    years = today.year - hire.year
    if (today.month, today.day) < (hire.month, hire.day):
        years -= 1
    return max(0, years)


def annual_days_for_tenure(n: int) -> float:
    if n < 1:
        return 0.0
    return float(min(25, 15 + (n - 1) // 2))


def is_annual_use(leave_type: str) -> bool:
    t = (leave_type or "").strip().lower()
    if not t:
        return False
    if "부여" in t:
        return False
    return "연차" in t or t in ("annual", "annual_leave")


def is_half(leave_type: str) -> bool:
    t = (leave_type or "").strip()
    return t.startswith("반차") or t.endswith("반차") or t == "반차"


def leave_days(row: dict) -> float:
    lt = str(row.get("leave_type") or "")
    if is_half(lt):
        return 0.5
    if row.get("days") is not None:
        try:
            d = float(row["days"])
            if d > 0:
                return d
        except (TypeError, ValueError):
            pass
    start = parse_ymd(row.get("start_date"))
    end = parse_ymd(row.get("end_date") or row.get("start_date"))
    if not start:
        return 0.0
    if not end or end < start:
        return 1.0
    return float((end - start).days + 1)


def resolve_hire(s: dict) -> date | None:
    return parse_ymd(s.get("hire_date")) or parse_ymd(s.get("join_date")) or parse_ymd(s.get("joined_at"))


def desktop_path() -> Path:
    # OneDrive Desktop 우선
    home = Path.home()
    candidates = [
        home / "OneDrive" / "Desktop",
        home / "OneDrive" / "바탕 화면",
        home / "Desktop",
        home / "바탕 화면",
    ]
    for c in candidates:
        if c.is_dir():
            return c
    # Windows special folder via PowerShell
    try:
        r = subprocess.run(
            ["powershell", "-NoProfile", "-Command", "[Environment]::GetFolderPath('Desktop')"],
            capture_output=True,
            text=True,
        )
        p = Path((r.stdout or "").strip())
        if p.is_dir():
            return p
    except Exception:
        pass
    return home


def style_header(ws, row: int, cols: int):
    fill = PatternFill("solid", fgColor="1F4E79")
    font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = fill
        cell.font = font
        cell.alignment = align
    ws.row_dimensions[row].height = 32


def autosize(ws, max_width=28):
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        width = 8
        for cell in col[:80]:
            if cell.value is not None:
                width = max(width, min(max_width, len(str(cell.value)) + 2))
        ws.column_dimensions[letter].width = width


def main():
    print("D1 조회 중…")
    staff = d1(
        "SELECT s.id, s.name, s.employee_no, s.company, s.department, s.team, s.status, "
        "s.hire_date, s.join_date, s.joined_at, s.annual_leave_total, s.annual_leave_used, "
        "b.year AS bal_year, b.total_days, b.used_days, b.remaining_days, b.expired_days, "
        "b.compensated_days, b.expiry_date, b.expired_at "
        "FROM staff_members s "
        "LEFT JOIN leave_balances b ON b.staff_id = s.id AND b.year = 2026 "
        "ORDER BY COALESCE(s.hire_date, s.join_date, s.joined_at), s.name"
    )
    accruals = d1(
        "SELECT staff_id, kind, period_key, days, year, source_date, note "
        "FROM leave_accruals ORDER BY staff_id, kind, period_key"
    )
    requests = d1(
        "SELECT staff_id, leave_type, start_date, end_date, status, days "
        "FROM leave_requests WHERE status IN ('승인', 'approved') "
        "ORDER BY staff_id, start_date"
    )
    print(f"staff={len(staff)} accruals={len(accruals)} requests={len(requests)}")

    by_acc: dict[str, list] = {}
    for a in accruals:
        by_acc.setdefault(str(a["staff_id"]), []).append(a)

    by_req: dict[str, list] = {}
    for r in requests:
        by_req.setdefault(str(r["staff_id"]), []).append(r)

    rows_summary = []
    rows_accrual_detail = []
    rows_usage_detail = []
    rows_mismatch = []

    for s in staff:
        sid = str(s["id"])
        name = s.get("name") or ""
        status = (s.get("status") or "").strip()
        hire = resolve_hire(s)
        hire_str = hire.isoformat() if hire else ""
        ty = tenure_years(hire, TODAY) if hire else None

        # --- DB balance ---
        db_total = float(s["total_days"]) if s.get("total_days") is not None else None
        db_used = float(s["used_days"]) if s.get("used_days") is not None else None
        db_remain = float(s["remaining_days"]) if s.get("remaining_days") is not None else None
        db_expired = float(s.get("expired_days") or 0)
        db_comp = float(s.get("compensated_days") or 0)
        staff_total = float(s["annual_leave_total"]) if s.get("annual_leave_total") is not None else None
        staff_used = float(s["annual_leave_used"]) if s.get("annual_leave_used") is not None else None

        # --- Accruals ---
        accs = by_acc.get(sid, [])
        monthly = [a for a in accs if a.get("kind") == "monthly"]
        annuals = [a for a in accs if a.get("kind") == "annual"]
        monthly_sum = sum(float(a.get("days") or 0) for a in monthly)
        annual_sum_all = sum(float(a.get("days") or 0) for a in annuals)
        accrual_sum_all = monthly_sum + annual_sum_all  # 목록 단순합 (이중합산 가능)

        # SSOT granted: latest annual N days, else monthly sum
        latest_annual_days = 0.0
        latest_annual_key = ""
        if annuals:
            best_n = -1
            for a in annuals:
                n = int(re.sub(r"^annual:", "", str(a.get("period_key") or "0")) or 0)
                if n >= best_n:
                    best_n = n
                    latest_annual_days = float(a.get("days") or 0)
                    latest_annual_key = str(a.get("period_key") or "")
            ssot_granted = latest_annual_days
            grant_source = f"annual:{best_n}" if best_n >= 0 else "annual"
        elif monthly_sum > 0:
            ssot_granted = monthly_sum
            grant_source = "monthly_sum"
        elif staff_total and staff_total > 0:
            ssot_granted = staff_total
            grant_source = "staff_fallback"
        else:
            ssot_granted = 0.0
            grant_source = "zero"

        # Expected by hire date (policy)
        exp_monthly = 0
        exp_annual_keys = []
        exp_granted = 0.0
        if hire:
            ty_v = tenure_years(hire, TODAY)
            if ty_v >= 1:
                for n in range(1, ty_v + 1):
                    exp_annual_keys.append(f"annual:{n}")
                exp_granted = annual_days_for_tenure(ty_v)
                exp_monthly = 0  # 잔액 기준은 연차만
            else:
                # months completed k=1..11 where end <= today
                for k in range(1, 12):
                    end = add_months(hire, k)
                    if end <= TODAY:
                        exp_monthly += 1
                exp_granted = float(exp_monthly)

        # Actual annual keys present
        annual_keys = sorted(
            str(a.get("period_key") or "") for a in annuals
        )
        missing_annual = [k for k in exp_annual_keys if k not in set(annual_keys)]
        monthly_cnt = len(monthly)
        annual_cnt = len(annuals)

        # --- Usage ---
        reqs = by_req.get(sid, [])
        use_all = 0.0
        use_2026 = 0.0
        use_2025 = 0.0
        use_after_anniv = 0.0  # hire anniversary period if tenure>=1
        anniv1 = add_years(hire, 1) if hire else None

        for r in reqs:
            if not is_annual_use(str(r.get("leave_type") or "")):
                continue
            days = leave_days(r)
            use_all += days
            sd = parse_ymd(r.get("start_date"))
            if not sd:
                continue
            if sd.year == 2026:
                use_2026 += days
            if sd.year == 2025:
                use_2025 += days
            if anniv1 and sd >= anniv1:
                use_after_anniv += days

        # Remaining formulas
        remain_db = db_remain
        remain_ssot = max(0.0, round(ssot_granted - (db_used if db_used is not None else use_2026) - db_expired - db_comp, 2))
        remain_by_hire_period = None
        if hire and tenure_years(hire, TODAY) >= 1:
            # 입사 응당일 기준: 총부여(최신 연차) - 1년차 응당일 이후 사용
            remain_by_hire_period = max(0.0, round(latest_annual_days - use_after_anniv - db_expired - db_comp, 2))
        elif hire:
            remain_by_hire_period = max(0.0, round(ssot_granted - use_all - db_expired - db_comp, 2))

        # Flags
        flags = []
        if status and status not in ("재직", "재직중", "active", "Active"):
            flags.append("비재직")
        if not hire:
            flags.append("입사일없음")
        if missing_annual:
            flags.append("연차부여누락:" + ",".join(missing_annual))
        if hire and tenure_years(hire, TODAY) < 1 and monthly_cnt < exp_monthly:
            flags.append(f"월차부족:{monthly_cnt}<{exp_monthly}")
        if hire and tenure_years(hire, TODAY) < 1 and monthly_cnt > exp_monthly:
            flags.append(f"월차초과:{monthly_cnt}>{exp_monthly}")
        if db_total is not None and abs(db_total - ssot_granted) > 0.01:
            flags.append(f"DB총부여≠SSOT({db_total}≠{ssot_granted})")
        if db_used is not None and abs(db_used - use_2026) > 0.01:
            flags.append(f"DB사용≠2026원장({db_used}≠{use_2026})")
        if db_remain is not None and abs(db_remain - remain_ssot) > 0.01:
            flags.append(f"DB잔여≠재계산({db_remain}≠{remain_ssot})")
        if abs(accrual_sum_all - ssot_granted) > 0.01 and annual_cnt > 0 and monthly_cnt > 0:
            flags.append(f"발생목록합이중합산가능({accrual_sum_all})")

        summary = {
            "name": name,
            "employee_no": s.get("employee_no") or "",
            "company": s.get("company") or "",
            "department": s.get("department") or "",
            "team": s.get("team") or "",
            "status": status,
            "hire_date": hire_str,
            "tenure_years": ty if ty is not None else "",
            "staff_total": staff_total if staff_total is not None else "",
            "staff_used": staff_used if staff_used is not None else "",
            "db_total": db_total if db_total is not None else "",
            "db_used": db_used if db_used is not None else "",
            "db_remaining": db_remain if db_remain is not None else "",
            "db_expired": db_expired,
            "db_compensated": db_comp,
            "db_expiry_date": s.get("expiry_date") or "",
            "monthly_cnt": monthly_cnt,
            "monthly_sum": monthly_sum,
            "annual_cnt": annual_cnt,
            "annual_sum_all": annual_sum_all,
            "accrual_list_sum": accrual_sum_all,
            "ssot_granted": ssot_granted,
            "grant_source": grant_source,
            "latest_annual_key": latest_annual_key,
            "exp_monthly_by_hire": exp_monthly if hire and (ty or 0) < 1 else "",
            "exp_annual_by_hire": ",".join(exp_annual_keys) if exp_annual_keys else "",
            "exp_granted_by_hire": exp_granted if hire else "",
            "use_all_time": use_all,
            "use_2025": use_2025,
            "use_2026": use_2026,
            "use_after_1y_anniv": use_after_anniv if anniv1 else "",
            "remain_ssot_recalc": remain_ssot,
            "remain_hire_period": remain_by_hire_period if remain_by_hire_period is not None else "",
            "flags": " | ".join(flags),
        }
        rows_summary.append(summary)

        if flags:
            rows_mismatch.append(summary)

        for a in accs:
            rows_accrual_detail.append(
                {
                    "name": name,
                    "employee_no": s.get("employee_no") or "",
                    "hire_date": hire_str,
                    "kind": a.get("kind"),
                    "period_key": a.get("period_key"),
                    "days": a.get("days"),
                    "year": a.get("year"),
                    "source_date": a.get("source_date"),
                    "note": a.get("note") or "",
                }
            )

        for r in reqs:
            if not is_annual_use(str(r.get("leave_type") or "")) and "부여" not in str(r.get("leave_type") or ""):
                # keep grants too for audit? include all annual-related
                pass
            lt = str(r.get("leave_type") or "")
            if "연차" not in lt and "반차" not in lt and lt.lower() not in ("annual", "annual_leave", "half_leave"):
                continue
            rows_usage_detail.append(
                {
                    "name": name,
                    "employee_no": s.get("employee_no") or "",
                    "hire_date": hire_str,
                    "leave_type": lt,
                    "start_date": r.get("start_date"),
                    "end_date": r.get("end_date"),
                    "status": r.get("status"),
                    "days_db": r.get("days") if r.get("days") is not None else "",
                    "days_calc": leave_days(r) if "부여" not in lt else 0,
                    "counts_as_use": "N" if "부여" in lt else "Y",
                    "year": str(r.get("start_date") or "")[:4],
                }
            )

    # --- Excel ---
    wb = Workbook()

    # Sheet 0: 안내
    ws0 = wb.active
    ws0.title = "안내"
    notes = [
        ["연차 전 직원 감사 자료", ""],
        ["생성일시", datetime.now().strftime("%Y-%m-%d %H:%M")],
        ["기준일(TODAY)", TODAY.isoformat()],
        ["데이터 소스", "Cloudflare D1 pchos-d1 (remote)"],
        ["", ""],
        ["시트 설명", ""],
        ["1_종합", "직원별 DB 잔액 vs 발생(SSOT) vs 입사일 기대 vs 사용 집계"],
        ["2_발생내역", "leave_accruals 원장 전건 (월차+연차)"],
        ["3_사용내역", "leave_requests 승인 연차/반차 전건"],
        ["4_불일치의심", "flags 가 있는 직원만"],
        ["", ""],
        ["용어", ""],
        ["accrual_list_sum", "발생 목록 일수 단순 합 (월차+연차 → 1년차 이상이면 이중합산 착시 가능)"],
        ["ssot_granted", "잔액 기준 총 부여: 최신 annual:N 일수, 없으면 월차 합"],
        ["use_2026", "2026년 시작일 기준 승인 연차 사용 합 (연차(부여) 제외)"],
        ["use_after_1y_anniv", "입사 1년 응당일 이후 사용 (입사일 기준 잔여 참고)"],
        ["remain_ssot_recalc", "ssot_granted - db_used(or use_2026) - expired - compensated"],
        ["remain_hire_period", "1년 이상: 최신연차 - 응당일이후 사용 / 1년 미만: ssot - 전기간 사용"],
        ["", ""],
        ["참고", "김영대 예시: accrual_list_sum=26(월11+연15), ssot_granted=15, use_2026=7, db_remaining=8"],
    ]
    for r in notes:
        ws0.append(r)
    ws0["A1"].font = Font(name="Arial", bold=True, size=14, color="1F4E79")
    ws0.column_dimensions["A"].width = 28
    ws0.column_dimensions["B"].width = 90

    def write_sheet(title: str, headers: list[str], data: list[dict], keys: list[str]):
        ws = wb.create_sheet(title)
        ws.append(headers)
        style_header(ws, 1, len(headers))
        thin = Border(
            left=Side(style="thin", color="D0D0D0"),
            right=Side(style="thin", color="D0D0D0"),
            top=Side(style="thin", color="D0D0D0"),
            bottom=Side(style="thin", color="D0D0D0"),
        )
        warn_fill = PatternFill("solid", fgColor="FFF2CC")
        bad_fill = PatternFill("solid", fgColor="FCE4D6")
        for i, row in enumerate(data, start=2):
            for c, k in enumerate(keys, start=1):
                val = row.get(k, "")
                cell = ws.cell(row=i, column=c, value=val if val != "" else None)
                cell.font = Font(name="Arial", size=9)
                cell.border = thin
                cell.alignment = Alignment(vertical="center")
            flags = str(row.get("flags") or "")
            if flags:
                fill = bad_fill if ("누락" in flags or "≠" in flags or "부족" in flags) else warn_fill
                for c in range(1, len(keys) + 1):
                    ws.cell(row=i, column=c).fill = fill
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{max(1, len(data)+1)}"
        ws.freeze_panes = "C2"
        autosize(ws)
        return ws

    headers1 = [
        "이름", "사번", "회사", "부서", "팀", "상태", "입사일", "근속년수(만)",
        "staff.total", "staff.used",
        "DB총부여", "DB사용", "DB잔여", "DB소멸", "DB보상", "DB만료일",
        "월차건수", "월차합", "연차건수", "연차합(전부)", "발생목록단순합",
        "SSOT총부여", "부여출처", "최신연차키",
        "기대월차(입사)", "기대연차키(입사)", "기대총부여(입사)",
        "사용_전기간", "사용_2025", "사용_2026", "사용_1년응당이후",
        "잔여_SSOT재계산", "잔여_입사기간기준", "플래그",
    ]
    keys1 = [
        "name", "employee_no", "company", "department", "team", "status", "hire_date", "tenure_years",
        "staff_total", "staff_used",
        "db_total", "db_used", "db_remaining", "db_expired", "db_compensated", "db_expiry_date",
        "monthly_cnt", "monthly_sum", "annual_cnt", "annual_sum_all", "accrual_list_sum",
        "ssot_granted", "grant_source", "latest_annual_key",
        "exp_monthly_by_hire", "exp_annual_by_hire", "exp_granted_by_hire",
        "use_all_time", "use_2025", "use_2026", "use_after_1y_anniv",
        "remain_ssot_recalc", "remain_hire_period", "flags",
    ]
    write_sheet("1_종합", headers1, rows_summary, keys1)

    headers2 = ["이름", "사번", "입사일", "kind", "period_key", "days", "year", "source_date", "note"]
    keys2 = ["name", "employee_no", "hire_date", "kind", "period_key", "days", "year", "source_date", "note"]
    write_sheet("2_발생내역", headers2, rows_accrual_detail, keys2)

    headers3 = [
        "이름", "사번", "입사일", "유형", "시작일", "종료일", "상태",
        "days(DB)", "days(계산)", "사용차감", "연도",
    ]
    keys3 = [
        "name", "employee_no", "hire_date", "leave_type", "start_date", "end_date", "status",
        "days_db", "days_calc", "counts_as_use", "year",
    ]
    write_sheet("3_사용내역", headers3, rows_usage_detail, keys3)

    write_sheet("4_불일치의심", headers1, rows_mismatch, keys1)

    # 요약 KPI sheet
    wsK = wb.create_sheet("0_요약KPI", 1)
    active = [r for r in rows_summary if r["status"] in ("재직", "재직중", "active", "Active", "")]
    flagged = [r for r in rows_summary if r["flags"]]
    kpi = [
        ["항목", "값"],
        ["전체 직원 수", len(rows_summary)],
        ["재직(대략) 수", len(active)],
        ["플래그 있는 직원", len(flagged)],
        ["발생내역 행수", len(rows_accrual_detail)],
        ["사용내역 행수", len(rows_usage_detail)],
        ["DB잔여 합(전직원)", round(sum(float(r["db_remaining"] or 0) for r in rows_summary), 2)],
        ["SSOT부여 합", round(sum(float(r["ssot_granted"] or 0) for r in rows_summary), 2)],
        ["2026사용 합", round(sum(float(r["use_2026"] or 0) for r in rows_summary), 2)],
        ["", ""],
        ["주의", "발생목록단순합은 월차+연차를 모두 더하므로 1년차 이상은 SSOT총부여와 다를 수 있음"],
    ]
    for row in kpi:
        wsK.append(row)
    style_header(wsK, 1, 2)
    wsK.column_dimensions["A"].width = 28
    wsK.column_dimensions["B"].width = 70

    out_dir = desktop_path()
    out_path = out_dir / f"연차_전직원_감사_{TODAY.isoformat()}.xlsx"
    wb.save(out_path)
    print(f"saved: {out_path}")
    print(f"flagged: {len(flagged)} / total: {len(rows_summary)}")
    # top flags preview
    for r in flagged[:15]:
        print(f"  - {r['name']}: {r['flags'][:120]}")


if __name__ == "__main__":
    main()
